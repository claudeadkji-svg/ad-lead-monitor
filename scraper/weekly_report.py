# -*- coding: utf-8 -*-
"""주간 리포트: 지난 7일간 수집된 전체 리드를 엑셀로 정리

매주 월요일 아침 GitHub Actions가 실행 → docs/reports/에 저장.
대시보드 하단에서 다운로드 링크 제공.
"""
import json
import os
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "docs", "data")
REPORTS = os.path.join(ROOT, "docs", "reports")

HEADERS = ["최초등록일", "카테고리", "소스", "제목", "거래처", "링크", "이메일", "전화번호", "설명"]
WIDTHS = [12, 16, 14, 55, 20, 45, 26, 15, 45]


def main():
    kst = timezone(timedelta(hours=9))
    now = datetime.now(kst)
    days = [(now - timedelta(days=d)).strftime("%Y-%m-%d") for d in range(7)]

    merged = {}
    for day in sorted(days):
        path = os.path.join(DATA, f"{day}.json")
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8") as f:
            snap = json.load(f)
        for it in snap.get("items", []):
            merged.setdefault(it["id"], it)

    items = sorted(merged.values(), key=lambda x: (x.get("first_seen", ""), x["category"]))
    if not items:
        print("지난 7일 데이터 없음 — 리포트 생략")
        return

    os.makedirs(REPORTS, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "주간 리드"
    ws.append(HEADERS)
    head_fill = PatternFill("solid", fgColor="A6192E")
    for c in ws[1]:
        c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for it in items:
        ws.append([it.get("first_seen", ""), it["category"], it["source"], it["title"],
                   it.get("company", ""), it.get("url", ""), it.get("email", ""),
                   it.get("phone", ""), it.get("description", "")])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name="맑은 고딕", size=10)
    ws.freeze_panes = "A2"

    week_label = f"{days[-1]}_{days[0]}"
    fname = f"weekly_{week_label}.xlsx"
    wb.save(os.path.join(REPORTS, fname))

    # 리포트 목록 갱신
    idx_path = os.path.join(REPORTS, "index.json")
    try:
        with open(idx_path, encoding="utf-8") as f:
            idx = json.load(f)
    except Exception:
        idx = {"reports": []}
    entry = {"file": fname, "label": f"{days[-1]} ~ {days[0]} ({len(items)}건)"}
    idx["reports"] = [r for r in idx["reports"] if r["file"] != fname]
    idx["reports"].insert(0, entry)
    idx["reports"] = idx["reports"][:52]
    with open(idx_path, "w", encoding="utf-8") as f:
        json.dump(idx, f, ensure_ascii=False, indent=1)

    print(f"주간 리포트 생성: {fname} ({len(items)}건)")


if __name__ == "__main__":
    main()
